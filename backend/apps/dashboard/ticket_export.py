from __future__ import annotations

import io
import re
from datetime import date, datetime, time, timedelta
from typing import Iterable

from django.contrib.contenttypes.models import ContentType
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from apps.activity.models import ActivityLog
from apps.projects.models import Project
from apps.tickets.models import Ticket

EXPORT_COLUMNS = [
    'Ticket ID',
    'Title',
    'Type',
    'Priority',
    'Status',
    'Module',
    'Project',
    'Created By',
    'Assignees',
    'Assigned By',
    'Created At',
    'Due Date',
    'Closed At',
    'Description',
]

DESCRIPTION_EXCEL_MAX = 500


def _escape_pdf_text(text: str) -> str:
    """Escape text for ReportLab Paragraph (minimal HTML)."""
    from xml.sax.saxutils import escape

    return escape((text or '').replace('\r\n', '\n').replace('\r', '\n'))


def _pdf_paragraph(text: str, style):
    from reportlab.platypus import Paragraph

    return Paragraph(_escape_pdf_text(text).replace('\n', '<br/>'), style)


def _display_name(user) -> str:
    if user is None:
        return ''
    full = f'{user.first_name or ""} {user.last_name or ""}'.strip()
    return full or user.username


def _normalize_period(period: str | None) -> str:
    value = (period or '30').strip().lower()
    if value in {'today', '7', '30', 'custom'}:
        return value
    return '30'


def resolve_export_date_range(
    period: str | None,
    start_date: str | None,
    end_date: str | None,
) -> tuple[date, date]:
    today = timezone.localdate()
    normalized = _normalize_period(period)

    if normalized == 'today':
        return today, today

    if normalized == '7':
        return today - timedelta(days=6), today

    if normalized == '30':
        return today - timedelta(days=29), today

    if not start_date or not end_date:
        raise ValueError('start_date and end_date are required for custom period.')

    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise ValueError('Invalid date format. Use YYYY-MM-DD.')

    if start > end:
        raise ValueError('start_date must be on or before end_date.')

    return start, end


def get_tickets_for_export(project_id: int, start: date, end: date) -> list[Ticket]:
    start_dt = timezone.make_aware(datetime.combine(start, time.min))
    end_dt = timezone.make_aware(datetime.combine(end, time.max))

    return list(
        Ticket.objects.filter(
            project_id=project_id,
            created_at__gte=start_dt,
            created_at__lte=end_dt,
        )
        .select_related('project', 'created_by')
        .prefetch_related('assignees')
        .order_by('created_at', 'id')
    )


def _assignment_actor_map(ticket_ids: Iterable[int]) -> dict[int, str]:
    if not ticket_ids:
        return {}

    ticket_ct = ContentType.objects.get_for_model(Ticket)
    logs = (
        ActivityLog.objects.filter(
            content_type=ticket_ct,
            object_id__in=ticket_ids,
        )
        .filter(
            Q(action='assignment_change')
            | Q(description__icontains='Assigned ticket')
            | Q(description__icontains='Self-assigned ticket')
        )
        .select_related('user')
        .order_by('object_id', '-created_at')
    )

    actors: dict[int, str] = {}
    for log in logs:
        if log.object_id in actors:
            continue
        if log.description and 'Self-assigned' in log.description:
            actors[log.object_id] = _display_name(log.user) or 'Self-assigned'
        else:
            actors[log.object_id] = _display_name(log.user) or '—'
    return actors


def _format_dt(value) -> str:
    if not value:
        return ''
    return timezone.localtime(value).strftime('%Y-%m-%d %H:%M')


def _format_date(value) -> str:
    if not value:
        return ''
    return value.isoformat()


def _truncate(text: str, limit: int) -> str:
    cleaned = re.sub(r'\s+', ' ', (text or '').strip())
    if len(cleaned) <= limit:
        return cleaned
    return cleaned[: limit - 1] + '…'


def build_export_rows(tickets: list[Ticket]) -> list[list[str]]:
    ticket_ids = [ticket.id for ticket in tickets]
    assignment_actors = _assignment_actor_map(ticket_ids)

    rows: list[list[str]] = []
    for ticket in tickets:
        assignees = ', '.join(
            _display_name(user) for user in ticket.assignees.all()
        ) or 'Unassigned'

        assigned_by = assignment_actors.get(ticket.id, '')
        if not assigned_by and ticket.assignees.exists():
            assigned_by = _display_name(ticket.created_by)

        rows.append([
            ticket.ticket_id,
            ticket.title,
            ticket.get_type_display(),
            ticket.get_priority_display(),
            ticket.get_status_display(),
            ticket.module or '',
            ticket.project.name,
            _display_name(ticket.created_by),
            assignees,
            assigned_by or '—',
            _format_dt(ticket.created_at),
            _format_date(ticket.due_date),
            _format_dt(ticket.closed_at),
            ticket.description or '',
        ])
    return rows


def _safe_filename(value: str) -> str:
    return re.sub(r'[^\w\-]+', '_', value).strip('_') or 'export'


def build_export_filename(project: Project, start: date, end: date, extension: str) -> str:
    project_slug = _safe_filename(project.name)
    return f'tickets_{project_slug}_{start.isoformat()}_to_{end.isoformat()}.{extension}'


def build_excel_response(
    project: Project,
    start: date,
    end: date,
    rows: list[list[str]],
) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Tickets'

    header_font = Font(bold=True)
    worksheet.append(EXPORT_COLUMNS)
    for cell in worksheet[1]:
        cell.font = header_font

    for row in rows:
        excel_row = row[:]
        if excel_row:
            excel_row[-1] = _truncate(excel_row[-1], DESCRIPTION_EXCEL_MAX)
        worksheet.append(excel_row)

    for idx, column_name in enumerate(EXPORT_COLUMNS, start=1):
        if column_name == 'Description':
            worksheet.column_dimensions[get_column_letter(idx)].width = 48
        elif column_name == 'Title':
            worksheet.column_dimensions[get_column_letter(idx)].width = 32
        else:
            worksheet.column_dimensions[get_column_letter(idx)].width = max(len(column_name) + 2, 14)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{build_export_filename(project, start, end, "xlsx")}"'
    )
    return response


def build_pdf_response(
    project: Project,
    start: date,
    end: date,
    rows: list[list[str]],
) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    page_width, _ = A4
    margin = 0.55 * inch
    content_width = page_width - (2 * margin)

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ExportTitle',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        spaceAfter=4,
        textColor=colors.HexColor('#0f172a'),
    )
    meta_style = ParagraphStyle(
        'ExportMeta',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        spaceAfter=10,
        textColor=colors.HexColor('#475569'),
    )
    ticket_title_style = ParagraphStyle(
        'TicketTitle',
        parent=styles['Heading3'],
        fontSize=10,
        leading=13,
        spaceAfter=6,
        textColor=colors.HexColor('#1e293b'),
    )
    label_style = ParagraphStyle(
        'FieldLabel',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#64748b'),
        fontName='Helvetica-Bold',
    )
    value_style = ParagraphStyle(
        'FieldValue',
        parent=styles['Normal'],
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#0f172a'),
    )
    description_style = ParagraphStyle(
        'Description',
        parent=styles['Normal'],
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#334155'),
        alignment=TA_LEFT,
        spaceBefore=2,
        spaceAfter=2,
    )

    story = [
        Paragraph('Ticket Export', title_style),
        Paragraph(
            f'<b>{_escape_pdf_text(project.name)}</b><br/>'
            f'Period: {start.isoformat()} to {end.isoformat()} · '
            f'{len(rows)} ticket(s)',
            meta_style,
        ),
    ]

    if not rows:
        story.append(Paragraph('No tickets found for the selected filters.', value_style))
    else:
        label_col = 0.95 * inch
        value_col = (content_width - (2 * label_col)) / 2
        meta_col_widths = [label_col, value_col, label_col, value_col]

        meta_table_style = TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f1f5f9')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
            ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#64748b')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
        ])

        for row in rows:
            ticket_id, title, ticket_type, priority, status, module, _project, created_by, assignees, assigned_by, created_at, due_date, closed_at, description = row

            meta_table = Table(
                [
                    [
                        _pdf_paragraph('Type', label_style),
                        _pdf_paragraph(ticket_type, value_style),
                        _pdf_paragraph('Priority', label_style),
                        _pdf_paragraph(priority, value_style),
                    ],
                    [
                        _pdf_paragraph('Status', label_style),
                        _pdf_paragraph(status, value_style),
                        _pdf_paragraph('Module', label_style),
                        _pdf_paragraph(module or '—', value_style),
                    ],
                    [
                        _pdf_paragraph('Created By', label_style),
                        _pdf_paragraph(created_by, value_style),
                        _pdf_paragraph('Assignees', label_style),
                        _pdf_paragraph(assignees, value_style),
                    ],
                    [
                        _pdf_paragraph('Assigned By', label_style),
                        _pdf_paragraph(assigned_by, value_style),
                        _pdf_paragraph('Created At', label_style),
                        _pdf_paragraph(created_at, value_style),
                    ],
                    [
                        _pdf_paragraph('Due Date', label_style),
                        _pdf_paragraph(due_date or '—', value_style),
                        _pdf_paragraph('Closed At', label_style),
                        _pdf_paragraph(closed_at or '—', value_style),
                    ],
                ],
                colWidths=meta_col_widths,
            )
            meta_table.setStyle(meta_table_style)

            description_block = Table(
                [
                    [_pdf_paragraph('Description', label_style)],
                    [_pdf_paragraph(description or '—', description_style)],
                ],
                colWidths=[content_width],
            )
            description_block.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))

            ticket_block = KeepTogether([
                Spacer(1, 8),
                _pdf_paragraph(f'{ticket_id} — {title}', ticket_title_style),
                meta_table,
                Spacer(1, 4),
                description_block,
                Spacer(1, 10),
            ])
            story.append(ticket_block)

    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="{build_export_filename(project, start, end, "pdf")}"'
    )
    return response


def build_ticket_export_response(
    project: Project,
    period: str | None,
    start_date: str | None,
    end_date: str | None,
    export_format: str,
) -> HttpResponse:
    """Build Excel or PDF export for a project's tickets within a date range."""
    normalized_format = (export_format or '').strip().lower()
    if normalized_format not in {'xlsx', 'pdf'}:
        raise ValueError('format must be xlsx or pdf.')

    start, end = resolve_export_date_range(period, start_date, end_date)
    tickets = get_tickets_for_export(project.id, start, end)
    rows = build_export_rows(tickets)

    if normalized_format == 'xlsx':
        return build_excel_response(project, start, end, rows)
    return build_pdf_response(project, start, end, rows)
